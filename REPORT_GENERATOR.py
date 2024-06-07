from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate
from reportlab.lib.colors import HexColor
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, Table, TableStyle
import pandas as pd
import subprocess
import sys
import os
import openpyxl
import matplotlib.pyplot as plt
from datetime import datetime
from dateutil.relativedelta import relativedelta
import yfinance as yf
import pandas as pd
from prophet import Prophet
from datetime import datetime
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def add_timestamp(c):
    timestamp = datetime.now().strftime('%d-%m-%Y %H:00')
    timestamp_style = ParagraphStyle(name='Timestamp', fontSize=7, fontName="Helvetica", alignment=0, textColor=HexColor("#000000"))
    timestamp_title = Paragraph(timestamp, timestamp_style)
    timestamp_title.wrapOn(c, 100, 50)
    timestamp_title.drawOn(c, 500, 20)

def create_pdf(data, ticker_symbol, filename="Company_Profile.pdf"):
    
    c = canvas.Canvas(filename, pagesize=letter)

    # Title
    title_style = ParagraphStyle(name='Title', fontSize=25, fontName="Helvetica-Bold", alignment=0, textColor=HexColor("#000000"))
    title_style2 = ParagraphStyle(name='Title', fontSize=25, fontName="Helvetica-Bold", alignment=1, textColor=HexColor("#324b00"))
    co_name = data.at[0, 'CO_Name'].split(" ")[0] + " " + data.at[0, "CO_Name"].split(" ")[1]
    title = Paragraph(co_name, title_style)
    title2 = Paragraph(data.at[0, "CO_Name"].split(" ")[2], title_style2)
    if len(co_name) <= 14:
        title.wrapOn(c, 200, 50)
        title.drawOn(c, 50, 750)
        title2.wrapOn(c, 300, 50)
        title2.drawOn(c, 80, 750)
    else:
        title.wrapOn(c, 300, 200)
        title.drawOn(c, 50, 750)
        title2.wrapOn(c, 300, 50)
        title2.drawOn(c, 180, 750)
    

    #Company Address
    address_style = ParagraphStyle(name='Address', fontSize=12, fontName="Helvetica", alignment=0, textColor=HexColor("#000000"), leading = 18)
    address_text = "<br/>".join([data.at[0, 'Location'], data.at[0, 'City'], data.at[0, 'Country']])
    address = Paragraph(address_text, address_style)
    address.wrapOn(c, 230, 100)
    address.drawOn(c, 50, 670)
    
    #Other Information
    other_style = ParagraphStyle(name='Other', fontSize=12, fontName="Helvetica", alignment=0, textColor=HexColor("#000000"), leading = 18)
    other_text = f"<strong>Sector:</strong> {data.at[0, 'Sector']}<br/><b>Industry:</b> {data.at[0, 'Industry']}<br/><b>Full Time Employees:</b> {data.at[0, 'Full Time Employees']}"
    other_info = Paragraph(other_text, other_style)
    other_info.wrapOn(c, 200, 100)
    other_info.drawOn(c, 300, 670)

    #Contact Info
    contact_style = ParagraphStyle(name = "Contact", fontSize = 10, fontName = "Helvetica-Bold", aligment = 0, textColor = HexColor("#324b00"))
    contact_text = f"{data.at[0, 'Phone Number']}<br/>{data.at[0, 'Website']}"
    contact_info = Paragraph(contact_text, contact_style)
    contact_info.wrapOn(c, 400, 20)
    contact_info.drawOn(c, 50, 630)

    #Upcoming Events
    ue_style = ParagraphStyle(name = "Upcoming Events", fontSize = 20, fontName = "Helvetica-Bold", aligment = 0, textColor = HexColor("#324b00"))
    ue_text = "Upcoming Events"
    ue_title = Paragraph(ue_text, ue_style)
    ue_title.wrapOn(c, 400, 30)
    ue_title.drawOn(c, 50, 560)

    e_dates_style = ParagraphStyle(name = "Event Dates", fontSize = 12, fontName="Helvetica-Bold", alignment=0, textColor=HexColor("#000000"), leading = 18)
    e_desc_style = ParagraphStyle(name = "Event Dates", fontSize = 12, fontName="Helvetica", alignment=0, textColor=HexColor("#000000"), leading = 18)
    events = data.at[0, 'UpcomingEvents'].split("NEW")
    k = 520
    for i in events:
        parts = i.split("\n")
        event_date = Paragraph(parts[0], e_dates_style)
        event_desc = Paragraph(parts[1], e_desc_style)
        event_date.wrapOn(c, 500, 20)
        event_desc.wrapOn(c, 500, 20)
        event_date.drawOn(c, 50, k)
        event_desc.drawOn(c, 50, k - 15)
        k = k - 50

    #Recent Events
    re_style = ParagraphStyle(name = "Recent Events", fontSize = 20, fontName = "Helvetica-Bold", aligment = 0, textColor = HexColor("#324b00"))
    re_text = "Recent Events"
    re_title = Paragraph(re_text, re_style)
    re_title.wrapOn(c, 400, 30)
    re_title.drawOn(c, 50, k - 10)

    events = data.at[0, 'RecentEvents'].split("NEW")
    k = k - 50
    for i in events:
        parts = i.split("\n")
        event_date = Paragraph(parts[0], e_dates_style)
        event_desc = Paragraph(parts[1], e_desc_style)
        event_date.wrapOn(c, 500, 20)
        event_desc.wrapOn(c, 500, 0)
        event_date.drawOn(c, 50, k)
        if len(parts[1]) < 90:
            event_desc.drawOn(c, 50, k - 15)
            k = k - 50
        else:
            event_desc.drawOn(c, 50, k - 55)
            k = k - 70

    data = pd.read_excel("C:/Users/Piotrek/Desktop/Uczelnia/III rok/Semestr II/Usługi Sieciowe w Biznesie/Projekt/Estimate.xlsx")
    wb = openpyxl.load_workbook("C:/Users/Piotrek/Desktop/Uczelnia/III rok/Semestr II/Usługi Sieciowe w Biznesie/Projekt/Estimate.xlsx")
    sheet_name = wb.sheetnames[0]
    wb.close()

    #data_values = [data.columns[:, None]] + data.values.tolist()
    data_values = [data.columns.to_numpy()[:, None]] + data.values.tolist()

    table = Table(data_values)

    table_title = Paragraph("Estimate for " + sheet_name, ue_style)
    table_title.wrapOn(c, 400, 30)
    table_title.drawOn(c, 50, k - 40)

    tablestyle = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), HexColor("#324b00")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('LEADING', (0, 0), (-1, -1), 14),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
    ])

    table.setStyle(tablestyle)
    table.wrapOn(c, 500, 500)
    table.drawOn(c, 100, k - 150)

    add_timestamp(c)

    c.showPage()

    min_date, max_date = last_time_chart("C:/Users/Piotrek/Desktop/Uczelnia/III rok/Semestr II/Usługi Sieciowe w Biznesie/Projekt/WEBSCRAPPING/stocks_history.xlsx", ticker_symbol)


    chart2_title_text = f"{co_name} Stock Price from {str(min_date)} to {str(max_date)}"
    chart_title = Paragraph(chart2_title_text, ue_style)
    chart_title.wrapOn(c, 500, 80)
    chart_title.drawOn(c, 50, 750)

    c.drawImage(f"C:/Users/Piotrek/Desktop/Uczelnia/III rok/Semestr II/Usługi Sieciowe w Biznesie/Projekt/{ticker_symbol}_stock_price_last_100.png", 50, 480, width = 500, height=220)

    
    csv_data = 'C:/Users/Piotrek/Desktop/Uczelnia/III rok/Semestr II/Usługi Sieciowe w Biznesie/Projekt/WEBSCRAPPING/stocks.csv'
    
    csv_text = add_csv_to_pdf(c, ticker_symbol, csv_data)

    if csv_text:
        # Add CSV data to the PDF
        csv_style = ParagraphStyle(name='CSV', fontSize=12, fontName="Helvetica", textColor=HexColor("#000000"), leading=18)
        csv_paragraph = Paragraph(csv_text, csv_style)
        csv_paragraph.wrapOn(c, 500, 800)
        csv_paragraph.drawOn(c, 50, 250)

    add_timestamp(c)

    c.showPage()

    chart_title_text = f"{co_name} Stock Price Over Last Year"
    chart_title = Paragraph(chart_title_text, ue_style)
    chart_title.wrapOn(c, 400, 80)
    chart_title.drawOn(c, 50, 750)

    forecast_file, timestamp_history = history_chart(ticker_symbol)

    c.drawImage(f"C:/Users/Piotrek/Desktop/Uczelnia/III rok/Semestr II/Usługi Sieciowe w Biznesie/Projekt/{ticker_symbol}_stock_price.png", 30, 500, width=500, height=220)
    
    forecast_data = append_prediction_topdf(forecast_file, timestamp_history, ticker_symbol)

    buy_count = (forecast_data['Indicator'] == 'Buy').sum()
    sell_count = (forecast_data['Indicator'] == 'Sell').sum()

    if buy_count > sell_count:
        summary = "Buy"
    else:
        summary = "Sell"

    table_data = [forecast_data.columns.values.tolist()] + forecast_data.values.tolist()
    forecast_table = Table(table_data, colWidths=[1.75 * inch, 1.75 * inch])

    prediction_title_text = "Predicted Prices for next week"
    prediction_title = Paragraph(prediction_title_text, ue_style)
    prediction_title.wrapOn(c, 400, 80)
    prediction_title.drawOn(c, 50, 420)

    tablestyle = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), HexColor("#324b00")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('LEADING', (0, 0), (-1, -1), 14),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
    ])

    forecast_table.setStyle(tablestyle)
    forecast_table.wrapOn(c, 500, 80)
    forecast_table.drawOn(c, 50, 250)

    summary_text = "Overall summary (all week):"
    summary_style = ParagraphStyle(name = "Week Summary", fontSize = 15, fontName = "Helvetica-Bold", aligment = 0, textColor = HexColor("#000000"))
    summary_title = Paragraph(summary_text, summary_style)
    summary_title.wrapOn(c, 300, 80)
    summary_title.drawOn(c, 310, 200)

    indicator_text = summary
    indicator_style = ParagraphStyle(name = "Week Summary", fontSize = 15, fontName = "Helvetica-Bold", aligment = 0, textColor = HexColor("#324b00"))
    indicator_text_with_underline = '<u>{}</u>'.format(indicator_text)
    indicator_title = Paragraph(indicator_text_with_underline, indicator_style)
    indicator_title.wrapOn(c, 150, 80)
    indicator_title.drawOn(c, 520 , 200)

    add_timestamp(c)

    c.showPage()

    today = datetime.today().strftime('%Y-%m-%d')
    table2_title = Paragraph(f"25 most active stocks - [{today}]", ue_style)
    table2_title.wrapOn(c, 400, 30)
    table2_title.drawOn(c, 50, 750)

    csv_25_data = 'C:/Users/Piotrek/Desktop/Uczelnia/III rok/Semestr II/Usługi Sieciowe w Biznesie/Projekt/WEBSCRAPPING/25_most_active_stocks.csv'
    add_csv_table_to_pdf(c, csv_25_data)

    add_timestamp(c)

    c.showPage()

    c.save()

    os.remove(f'C:/Users/Piotrek/Desktop/Uczelnia/III rok/Semestr II/Usługi Sieciowe w Biznesie/Projekt/{ticker_symbol}_stock_price.png')
    os.remove(f'C:/Users/Piotrek/Desktop/Uczelnia/III rok/Semestr II/Usługi Sieciowe w Biznesie/Projekt/{ticker_symbol}_stock_price_last_100.png')

def add_csv_table_to_pdf(canvas, stocks_csv_path):
    # Load the stocks data CSV
    stocks_data = pd.read_csv(stocks_csv_path)
    
    stocks_data.rename(columns={"Avg Vol (3 month)": "Avg Vol"}, inplace=True)
    stocks_data.rename(columns={"Price (Intraday)": "Price"}, inplace=True)

    stocks_data.fillna("---", inplace=True)

    # Prepare the data for the table
    table_data = [stocks_data.columns.tolist()] + stocks_data.values.tolist()
    
    # Create the table
    table = Table(table_data, colWidths=[0.65 * inch, 1.3 * inch, 0.7 * inch, 0.7 * inch, 0.7 * inch, 0.7 * inch, 0.7 * inch, 0.9 * inch, 0.9 * inch])
    
    # Add style to the table
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#324b00")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),
        ('LINEABOVE', (0, 1), (-1, -1), 0.5, colors.black),
        ('LINEBELOW', (0, -1), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    # Draw the table on the canvas
    table.wrapOn(canvas, 0, 0)
    table.drawOn(canvas, 0.5 * inch, 3.5 * inch)

def add_csv_to_pdf(canvas, ticker_symbol, csv_path):
    # Load the CSV data
    data = pd.read_csv(csv_path)
    
    # Filter data for the specific ticker symbol
    company_data = data[data['Symbol'] == ticker_symbol]
    
    if company_data.empty:
        return

    # Extract column names and values
    col_names = company_data.columns.tolist()
    col_values = company_data.values.tolist()[0]
    
    # Prepare the PDF elements
    styles = getSampleStyleSheet()
    header_style = ParagraphStyle(name='HeaderStyle', parent=styles['Heading2'], textColor=colors.HexColor("#324b00"), fontName='Helvetica-Bold')
    value_style = ParagraphStyle(name='ValueStyle', parent=styles['BodyText'], fontName='Helvetica')
    
    elements = []
    
    for col_name, col_value in zip(col_names, col_values):
        # Format the column name
        formatted_col_name = col_name.replace('_', ' ').title()
        # Create the header and value paragraphs
        header_paragraph = Paragraph(formatted_col_name, header_style)
        value_paragraph = Paragraph(str(col_value), value_style)
        # Add the paragraphs to elements
        elements.append([header_paragraph, value_paragraph])
    
    # Create a table with the elements
    table = Table(elements, colWidths=[3.5 * inch, 3.5 * inch])
    
    # Add table style
    table.setStyle(TableStyle([
        ('LINEABOVE', (0, 0), (-1, -1), 0.5, colors.black),
        ('LINEBELOW', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor("#324b00")),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
    ]))
    
    # Draw the table on the canvas, adjust the position and size
    width, height = letter
    table.wrapOn(canvas, width, height)
    table.drawOn(canvas, 0.7 * inch, height - 9.5 * inch)

'''
def add_csv_to_pdf(canvas, ticker_symbol, company_csv_path, stocks_csv_path):
    # Load the company data CSV
    company_data = pd.read_csv(company_csv_path)
    
    # Filter data for the specific ticker symbol
    filtered_company_data = company_data[company_data['Symbol'] == ticker_symbol]
    
    if filtered_company_data.empty:
        return

    # Extract column names and values
    col_names = filtered_company_data.columns.tolist()
    col_values = filtered_company_data.values.tolist()[0]
    
    # Prepare the PDF elements for the company data
    styles = getSampleStyleSheet()
    header_style = ParagraphStyle(name='HeaderStyle', parent=styles['Heading2'], textColor=colors.HexColor("#324b00"), fontName='Helvetica-Bold')
    value_style = ParagraphStyle(name='ValueStyle', parent=styles['BodyText'], fontName='Helvetica')
    
    elements = []
    
    for col_name, col_value in zip(col_names, col_values):
        # Format the column name
        formatted_col_name = col_name.replace('_', ' ').title()
        # Create the header and value paragraphs
        header_paragraph = Paragraph(formatted_col_name, header_style)
        value_paragraph = Paragraph(str(col_value), value_style)
        # Add the paragraphs to elements
        elements.append([header_paragraph, value_paragraph])
    
    # Create a table with the elements
    table1 = Table(elements, colWidths=[2.8 * inch, 2.8 * inch])
    
    # Add table style
    table1.setStyle(TableStyle([
        ('LINEABOVE', (0, 0), (-1, -1), 0.5, colors.black),
        ('LINEBELOW', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor("#324b00")),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
    ]))
    
    # Load the stocks data CSV
    stocks_data = pd.read_csv(stocks_csv_path)
    
    # Prepare the data for the second table
    second_table_data = [stocks_data.columns.tolist()] + stocks_data.values.tolist()
    
    # Create the second table
    table2 = Table(second_table_data, colWidths=[0.8 * inch, 1.2 * inch, 1.2 * inch, 0.8 * inch, 0.8 * inch, 1.0 * inch, 1.0 * inch, 1.2 * inch, 1.2 * inch])
    
    # Add style to the second table
    table2.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#324b00")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),
        ('LINEABOVE', (0, 1), (-1, -1), 0.5, colors.black),
        ('LINEBELOW', (0, -1), (-1, -1), 1, colors.black),
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
    ]))
    
    # Draw the tables on the canvas
    table1.wrapOn(canvas, 0, 0)
    table1.drawOn(canvas, 0.5 * inch, 5 * inch)
    
    table2.wrapOn(canvas, 0, 0)
    table2.drawOn(canvas, 0.5 * inch, 0.5 * inch)
'''
    
def append_prediction_topdf(file, sheet, ticker_symbol):
    forecast_data = pd.read_excel(file, sheet_name=sheet)

    todays_date = datetime.today()
    year_ago = todays_date - relativedelta(years=1)
    date = year_ago.strftime('%Y-%m-%d')

    data = yf.download(ticker_symbol, start=date)

    todays_close = data['Close'].iloc[-1]

    forecast_data['MA'] = (forecast_data['yhat'] - todays_close) / todays_close * 100

    forecast_data['ds'] = pd.to_datetime(forecast_data['ds']).dt.strftime('%Y-%m-%d')

    forecast_data.rename(columns={
        'yhat': 'Predicted Price',
        'Indicator': 'Indicator',
        'ds': 'Date',
        'MA': 'Change (percent)'
    }, inplace=True)
    #print(forecast_data.head())

    forecast_data['Change (percent)'] = forecast_data['Change (percent)'].apply(lambda x: f"+{x:.2f}%" if x > 0 else f"{x:.2f}%")

    return forecast_data


def append_to_forecast_history(filename, data, timestamp, ticker):
    if os.path.isfile(filename):
        # Plik istnieje, otwórz go w trybie append
        with pd.ExcelWriter(filename, mode="a", engine="openpyxl") as writer:
            sheet_name = timestamp
            data.to_excel(writer, index=False, sheet_name=sheet_name)
    else:
        # Plik nie istnieje, stwórz nowy plik z nazwą {ticker}_FORECAST.xlsx
        with pd.ExcelWriter(filename, mode="w", engine="openpyxl") as writer:
            sheet_name = timestamp
            data.to_excel(writer, index=False, sheet_name=sheet_name)

def history_chart(ticker_symbol):
    todays_date = datetime.today()
    year_ago = todays_date - relativedelta(years=1)
    date = year_ago.strftime('%Y-%m-%d')

    data = yf.download(ticker_symbol, start=date)
    data_close = data[['Close']]
    data_close.reset_index(inplace=True)
    data_close.columns = ['ds', 'y']

    # Prophet model training
    m = Prophet(yearly_seasonality=True)
    m.fit(data_close)
    # Forecast for the next 7 days
    future = m.make_future_dataframe(periods=7)
    forecast = m.predict(future)
    
    # Get the forecast for the next 7 days
    forecast_next_week = forecast[['ds', 'yhat']].tail(7)

    # Add moving average and indicator columns
    forecast['MA'] = forecast['yhat'].rolling(window=80).mean().fillna(0)
    forecast['Indicator'] = np.where(forecast['yhat'] < forecast['MA'], 'Buy', 'Sell')

    #forecast = forecast[forecast['ds'] >= date]

    # Plot historical data and forecast
    plt.figure(figsize=(15, 8))
    fig1 = m.plot(forecast)
    
    # Add highest, lowest, and today's close price
    highest_close = data['Close'].max()
    lowest_close = data['Close'].min()
    todays_close = data['Close'].iloc[-1]
    volume = data['Volume'].iloc[-1]
    
    #plt.axhline(highest_close, color='red', linestyle='--', label=f'Highest Close: {highest_close}')
    #plt.axhline(lowest_close, color='blue', linestyle='--', label=f'Lowest Close: {lowest_close}')
    #plt.axhline(todays_close, color='green', linestyle='--', label=f'Today\'s Close: {todays_close}')

    min_close_date = data['Close'].idxmin()
    max_close_date = data['Close'].idxmax() 
    today_date = data.index[-1]

    plt.scatter(min_close_date, lowest_close, color='red', marker='v', s=150, label='Min Close Price')
    plt.scatter(max_close_date, highest_close, color='green', marker='^', s=150, label='Max Close Price')
    plt.scatter(today_date, todays_close, color = "cyan", marker="o", s=100, label='Today')
    plt.legend()

    min_volume = data['Volume'].min()
    max_volume = data['Volume'].max()
    volume_text = f"Volume Range: {min_volume} - {max_volume}"
    plt.figtext(0.1, 0.02, volume_text, ha="left", fontsize=12, fontweight='bold')

    plt.xlabel('Date')
    plt.ylabel('Price')
    plt.title(f'{ticker_symbol} Stock Price with Prophet Forecast')
    plt.grid(True)

    # Save plot to file
    plt.savefig(f'C:/Users/Piotrek/Desktop/Uczelnia/III rok/Semestr II/Usługi Sieciowe w Biznesie/Projekt/{ticker_symbol}_stock_price.png')

    forecast_next_week['MA'] = forecast['yhat'].rolling(window=80).mean().fillna(0)


    # Save forecast to Excel file
    timestamp = datetime.now().strftime("%Y_%m_%d %H_%M_%S")
    filename = f'C:/Users/Piotrek/Desktop/Uczelnia/III rok/Semestr II/Usługi Sieciowe w Biznesie/Projekt/FORECAST/forecast_{ticker_symbol}.xlsx'
    forecast_next_week['Indicator'] = forecast_next_week.apply(lambda row: 'Buy' if row['yhat'] > row['MA'] else 'Sell', axis=1)
    #print(forecast_next_week.head())
    append_to_forecast_history(filename, forecast_next_week, timestamp, ticker_symbol)
    
    return filename, timestamp



def last_time_chart(file_path, ticker_symbol):
    sheet_name = ticker_symbol

    df = pd.read_excel(file_path, sheet_name=sheet_name)
    df['Datetime'] = pd.to_datetime(df['Date'].astype(str) + ' ' + df['Time'].astype(str), dayfirst=True)

    daily_prices_path = 'C:/Users/Piotrek/Desktop/Uczelnia/III rok/Semestr II/Usługi Sieciowe w Biznesie/Projekt/WEBSCRAPPING/AAPL.csv'
    daily_prices_df = pd.read_csv(daily_prices_path)
    daily_prices_df['Datetime'] = pd.to_datetime(daily_prices_df['Date'] + ' 00:00:00', dayfirst=True)

    #df_last_100 = df.tail(100)
    df_last_100 = df.iloc[4:122]

    done = False
    already_appended = False
    ready_df = pd.DataFrame()

    while not done:
        if df_last_100['regular_market_price'].iloc[-20:].nunique() == 1:
            if not already_appended:
                ready_df = df_last_100.tail(2)
                already_appended = True
            df_last_100 = df_last_100.iloc[:-40].tail(100)
        else:
            done = True
            ready_df = pd.concat([df_last_100, ready_df])

    min_date = ready_df['Datetime'].min().date()
    max_date = ready_df['Datetime'].max().date()

    all_dates = pd.date_range(start=min_date, end=max_date, freq='D')

    for date in all_dates:
        if date not in ready_df['Datetime'].dt.date.values:
            close_price = daily_prices_df[daily_prices_df['Datetime'].dt.date == date]['Close']
            if not close_price.empty:
                new_row = {
                    'Datetime': pd.Timestamp(date),
                    'Date': date.strftime('%d-%m-%Y'),
                    'regular_market_price': close_price.values[0]
                }
                ready_df = ready_df.append(new_row, ignore_index=True)

    ready_df = ready_df.sort_values(by='Datetime').reset_index(drop=True)

    chart_df = ready_df[['Datetime', 'Date', 'regular_market_price']]

    plt.figure(figsize=(10, 5))
    plt.plot(chart_df['Datetime'], chart_df['regular_market_price'], color='darkgreen', linewidth=1.5, label='_nolegend')

    plt.title('')
    plt.xlabel('Date')
    plt.ylabel('Price')
    plt.legend()
    plt.grid(True)

    #unique_dates = chart_df['Datetime'].dt.strftime('%Y-%m-%d').unique()
    unique_dates = chart_df['Datetime'].dt.date.unique()

    plt.xticks(pd.to_datetime(unique_dates))

    plt.gcf().autofmt_xdate()

    plt.savefig(f'C:/Users/Piotrek/Desktop/Uczelnia/III rok/Semestr II/Usługi Sieciowe w Biznesie/Projekt/{ticker_symbol}_stock_price_last_100.png')

    return min_date, max_date
    
def decorator():
    print("\n######################################################")


if len(sys.argv) <= 1:
    print('Ticker symbol CLI argument missing!')
    sys.exit(2)
elif len(sys.argv) > 2:
    print('Enter only one CLI argument!')
    sys.exit(2)
else:
    decorator()
    print(f"\nGENERATING REPORT FOR {sys.argv[1]} Please wait...")
    decorator()

    with open(os.devnull, 'w') as null: 
        subprocess.run(["python", "company.py", sys.argv[1]], stdout=null, stderr=null) 
        # Load data from CSV
        
    data = pd.read_csv('profile.csv')

        # Generate PDF
    create_pdf(data, sys.argv[1])
